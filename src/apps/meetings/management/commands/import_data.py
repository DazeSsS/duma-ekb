import re
import glob
from pathlib import Path
from openpyxl import load_workbook
from dateutil.parser import parse
from django.db import IntegrityError
from django.conf import settings
from django.core.management.base import BaseCommand

from apps.meetings.domain.enums import AuthorClassification, MeetingType, Solution
from apps.meetings.models import ImportedFile, Meeting, Question, Tag


COLUMN_ORDER = [
    'year',
    'month',
    'day',
    'protocol_number',
    'meeting_type',
    'deputies',
    'presiding',
    'quorum',
    'position_1870',
    'position_1892',
    'author_classification',
    'number',
    'description',
    'tags',
    'solution',
    'solution_content',
    'case_number',
    'sheet_numbers',
]

month_map = {
    'январь': 'Jan', 'февраль': 'Feb', 'март': 'Mar', 'апрель': 'Apr',
    'май': 'May', 'июнь': 'Jun', 'июль': 'Jul', 'август': 'Aug',
    'сентябрь': 'Sep', 'октябрь': 'Oct', 'ноябрь': 'Nov', 'декабрь': 'Dec'
}


class Command(BaseCommand):
    help = 'Мигрирует данные из Excel файлов'

    def get_columns(self, rows):
        header_row = [column for column in next(rows) if column is not None]
        if len(header_row) == 19:
            is_old = True
        elif len(header_row) == 18:
            is_old = False
        else:
            self.stdout.write(self.style.ERROR('Ошибка: неизвестная структура документа'))
            return

        columns = {}
        for idx, column in enumerate(header_row):
            if idx > 11 and is_old:
                columns[COLUMN_ORDER[idx - 1]] = idx
            else:
                columns[COLUMN_ORDER[idx]] = idx

        return columns

    def handle(self, *args, **options):
        files = [Path(file) for file in glob.glob(f'{settings.DATA_PATH}/*.xlsx')]
        imported_files = ImportedFile.objects.values_list('filename', flat=True)

        for file in files:
            if file.name in imported_files:
                choice = ''
                while choice not in ['1', '2']:
                    choice = input(
                        'Что вы хотите сделать с уже импортированными файлами?\n'
                        '1) Перезаписать данные из этих файлов\n'
                        '2) Пропустить\n'
                        '>> '
                    )
                break

        for file in files:
            imported_file = ImportedFile.objects.filter(filename=file.name)
            if imported_file.exists():
                if choice == '1':        
                    imported_file.delete()
                else:
                    continue

            workbook = load_workbook(file, read_only=True)
            first_sheet = workbook.sheetnames[0]
            sheet = workbook[first_sheet]

            rows = sheet.values
            columns = self.get_columns(rows)
            if columns is None:
                continue

            total_rows = 0
            imported_rows = 0

            source_file = ImportedFile.objects.create(filename=file.name)

            actual_meeting = None
            for idx, row in enumerate(rows):
                if not any(value is not None for value in row):
                    break

                total_rows += 1

                try:
                    date = parse(f'{row[columns['year']]} {month_map[row[columns['month']].lower().strip()]} {row[columns['day']]}').date()
                    protocol_number = (
                        (str(row[columns['protocol_number']]) or '')
                        .replace('(', '')
                        .replace(')', '')
                        .strip()
                    )
                    meeting_type = MeetingType.get_value_by_label(row[columns['meeting_type']].split()[0].strip())
                    deputies = int(row[columns['deputies']])
                    presiding = row[columns['presiding']].strip()
                    case_number = str(row[columns['case_number']] or '').strip()

                    if actual_meeting is None or actual_meeting.date != date:
                        new_meeting = Meeting.objects.create(
                            date=date,
                            protocol_number=protocol_number,
                            meeting_type=meeting_type,
                            deputies=deputies,
                            presiding=presiding,
                            case_number=case_number,
                            source_file_id=source_file.id,
                        )
                        actual_meeting = new_meeting

                    meeting_id = actual_meeting.id
                    number = (str(row[columns['number']]) or '').strip()
                    description = str(row[columns['description']] or '').strip()
                    quorum = True if row[columns['quorum']].strip().lower() == 'да' else False
                    position_1870 = row[columns['position_1870']][:1].capitalize().strip() if row[columns['position_1870']] else ''
                    position_1892 = row[columns['position_1892']].strip() if row[columns['position_1892']] else ''
                    author_classification = AuthorClassification.get_value_by_label(row[columns['author_classification']].strip())
                    solution = Solution.get_value_by_label(row[columns['solution']].strip()) if row[columns['solution']] else ''
                    solution_content = str(row[columns['solution_content']] or '').strip()
                    tags = row[columns['tags']] or ''

                    sheet_numbers = str(row[columns['sheet_numbers']]).split('-')
                    sheet_numbers_list = []

                    for part in sheet_numbers:
                        part = part.lower().strip().replace(' ', '').replace('.', '')
                        if 'об' in part:
                            num = part.replace('об', '.5')
                        else:
                            num = part

                        sheet_numbers_list.append(num)

                    if len(sheet_numbers_list) == 1:
                        sheet_numbers_list.append(sheet_numbers_list[0])

                    new_question = Question.objects.create(
                        meeting_id=meeting_id,
                        number=number,
                        description=description,
                        quorum=quorum,
                        position_1870=position_1870,
                        position_1892=position_1892,
                        author_classification=author_classification,
                        solution=solution,
                        solution_content=solution_content,
                        sheet_numbers=sheet_numbers_list,
                    )

                    for tag in re.split(r'[;,]+', tags):
                        cleaned_tag = tag.strip()
                        if cleaned_tag:
                            cleaned_tag = f'{cleaned_tag[0].capitalize()}{cleaned_tag[1:] if len(cleaned_tag) > 1 else ''}'
                            new_tag = Tag.objects.get_or_create(title=cleaned_tag)[0]
                            new_question.tags.add(new_tag)

                    imported_rows += 1

                except Exception as e:
                    self.stdout.write(self.style.ERROR(f'Не удалось обработать строку {idx + 2} в файле "{file}": {e}'))

            self.stdout.write(self.style.SUCCESS(f'{file}: {imported_rows}/{total_rows}'))
